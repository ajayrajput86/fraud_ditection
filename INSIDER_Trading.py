import codecs
import io
import logging
import os
import re
import shutil
import urllib
from datetime import date
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from zipfile import ZipFile

import numpy as np
import openpyxl
import pandas as pd
import requests
import xmltodict
from bs4 import BeautifulSoup
# urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side, PatternFill
from openpyxl.styles.numbers import builtin_format_code, BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from pandas.io import json

final_excel_file_name = "final.xlsx"
final_fii_cash_segment_file = "cashSegment.xlsx"
final_inside_trading_file = "InsiderTrading.xlsx"
root_path = os.getcwd() + "\\NSE\\"
final_app_it_data = ""
final_app_id_data = ""
support_path = root_path + "support\\"
backup_path = root_path + "\\backup\\"
final_excel_name = root_path + final_excel_file_name
final_CashSegment_file_name = root_path + final_fii_cash_segment_file
final_inside_trading_file_name = root_path + final_inside_trading_file
base_url = 'https://www1.nseindia.com'
indentation = "         "
support_debug = False
# Sheet Names
IndexReturnSetupSheet = "Index Return Setup"
IndexReturnSheet = "Index Return Daily"
IndexReturnSummarySheet = "Index Return 30D"
StocksSheet = "Stocks"
FNOSheet = "FNO"
TopValueSheet = "Top value"
ActiveSecuritiesValueSheet = "Most active security value"
Top25VolumeGainersSheet = "Top 25 - Volume Gainers"
ActiveSecuritiesVolumeSheet = "Most active security volume"
OISpurtsSheet = "OI Spurt"
NiftyTop10Sheet = "Nifty weight top 10"
MWPLSheet = "MWPL"

FIIDerivativeSheet = "FII Derivative"
CashSegmentSheet = "Cash Segment"
Participant_Interest_Sheet = "Participant Open Interest"
Participant_Volumes_Sheet = "Participant Trading Volumes"

# Columns
column_oiSpurtsDate = "Date"
column_oiSpurtsRank = "Rank"
column_oiSpurtsTotal = "volume"
column_ActiveSecurityDate = 'Date'
column_Top25VolumeGainersDate = 'Date'
column_Top25VolumeGainersVolume = 'Today Volume'
column_TopValueDate = 'Date'
column_EquitiesStockWatchSector = 'Sector'
column_TopValueRank = 'Rank'
column_CashSegmentDate = 'Date'

# heading Style
headingStyle = '40 % - Accent3'
topEmptyRows = 3
FORMAT_PERCENT_COLOR = u'##0.00;[Red]-##0.00;'
FORMAT_NUMBER_ONE_DECIMAL = u'#,##0.0;[Red]-#,##0.0;'

NSE_SELL_COL_NAME = 'Sell Value in Rs.'
NSE_BUY_COL_NAME = 'Buy Value in Rs.'
NSE_OTHERS = 'OTHERS'
NSE_PROTRADE = 'PRO-TRADES'
NSE_BANK = 'BNK'
NSE_CATEGORY = 'Category'
head_retail = "Retail + HNI (client)"
head_proprietary = "Proprietary"
head_bank_nri = "Bank + NRI + others"

column_participant_date = "Date"


def setup():
    try:
        print(root_path)
        Path(backup_path).mkdir(parents=True, exist_ok=True)
        Path(support_path).mkdir(parents=True, exist_ok=True)
        get_request_boot()
    except Exception as e:
        print("setup error " + e.strerror)


def take_backup():
    try:
        now = str(datetime.now())[:19]
        now = now.replace(":", "_")
        new_file_name = backup_path + str(now) + "-" + final_excel_file_name
        shutil.copy(final_excel_name, new_file_name)
        new_file_name = backup_path + str(now) + "-" + final_fii_cash_segment_file
        shutil.copy(final_CashSegment_file_name, new_file_name)
        new_file_name = backup_path + str(now) + "-" + final_inside_trading_file
        shutil.copy(final_inside_trading_file_name, new_file_name)
    except Exception as e:
        print(e)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    thin_border = Border(left=Side(style='thick'))

    try:
        # try to open an existing workbook
        writer.book = openpyxl.load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, index=False, startrow=startrow, **to_excel_kwargs)

    worksheet = writer.book[sheet_name]

    FullRange = "A4:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange

    MIN_WIDTH = 15
    for i, column_cells in enumerate(worksheet.columns, start=1):
        width = (length if (length := max(
            len(str(cell_value) if (cell_value := cell.value) is not None else "") for cell in
            column_cells)) >= MIN_WIDTH else MIN_WIDTH)
        worksheet.column_dimensions[get_column_letter(i)].width = width

    try:
        if "Reportheader" not in writer.book.named_styles:
            header = NamedStyle(name="Reportheader")
            header.font = Font(bold=True)
            header.border = Border(bottom=Side(border_style="thin"))
            header.alignment = Alignment(horizontal="center", vertical="center")
            writer.book.add_named_style(header)
        else:
            header = "Reportheader"
        if "dateformat" not in writer.book.named_styles:
            dateformat = NamedStyle(name="dateformat", number_format="DD-MMM-YY")
    except ValueError as e:
        logging.warning("style creation skipped because {}".format(e))

    header_row = worksheet[4]
    for cell in header_row:
        cell.style = headingStyle

    row_start = 5
    if sheet_name == Top25VolumeGainersSheet:

        for i in range(row_start, len(df) + row_start):
            for colNum in {4, 5, 6}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in {9, 10}:
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_NUMBER_ONE_DECIMAL
            for colNum in {7, 8, 11}:
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == IndexReturnSheet:
        for i in range(row_start, len(df) + row_start):
            # worksheet.cell(row=i, column=2).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            for colNum in {2, 3, 4}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(5, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == IndexReturnSummarySheet:
        for i in range(row_start, len(df) + row_start):
            worksheet.cell(row=i, column=2).number_format = FORMAT_PERCENT_COLOR
            for colNum in {3, 4}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(5, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == StocksSheet:
        for i in range(row_start, len(df) + row_start):
            worksheet.cell(row=i, column=3).number_format = FORMAT_PERCENT_COLOR
            for colNum in {4, 5}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(6, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == FNOSheet:
        for i in range(row_start, len(df) + row_start):
            worksheet.cell(row=i, column=3).number_format = FORMAT_PERCENT_COLOR
            for colNum in {4, 5, 6, 7}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(8, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == TopValueSheet:
        for i in range(row_start, len(df) + row_start):
            for colNum in {3, 5, 6}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            worksheet.cell(row=i, column=4).number_format = FORMAT_PERCENT_COLOR
            worksheet.cell(row=i, column=5).number_format = builtin_format_code(49)
            # FORMAT_TEXT

    if sheet_name == ActiveSecuritiesValueSheet:
        for i in range(row_start, len(df) + 3):
            worksheet.cell(row=i, column=4).number_format = BUILTIN_FORMATS[3]
            worksheet.cell(row=i, column=3).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == ActiveSecuritiesVolumeSheet:
        for i in range(row_start, len(df) + row_start):
            worksheet.cell(row=i, column=4).number_format = BUILTIN_FORMATS[3]
            worksheet.cell(row=i, column=3).number_format = FORMAT_PERCENT_COLOR

    if sheet_name == OISpurtsSheet:
        for i in range(row_start, len(df) + row_start):
            for colNum in {4, 5, 6, 7, 8}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(10, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR
            worksheet.cell(row=i, column=2).number_format = BUILTIN_FORMATS[38]
            worksheet.cell(row=i, column=3).number_format = BUILTIN_FORMATS[3]
            worksheet.cell(row=i, column=9).number_format = BUILTIN_FORMATS[3]
    if sheet_name == MWPLSheet:
        for i in range(row_start, len(df) + row_start):
            for colNum in {2, 3}:
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
            for colNum in range(4, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR
    if sheet_name == CashSegmentSheet:

        worksheet.freeze_panes = "B2"
        worksheet.row_dimensions[4].height = 50
        worksheet.merge_cells('B2:D2')
        worksheet.cell(row=2, column=2).value = 'FII'
        worksheet.cell(row=2, column=2).alignment = Alignment(horizontal='center')
        worksheet.cell(row=2, column=2).fill = PatternFill(start_color="FFFF00", fill_type="solid")
        worksheet.merge_cells('E2:G2')
        worksheet.cell(row=2, column=5).value = 'DII'
        worksheet.cell(row=2, column=5).alignment = Alignment(horizontal='center')
        worksheet.cell(row=2, column=5).fill = PatternFill(start_color="DAEEF3", fill_type="solid")

        worksheet.merge_cells('H2:J2')
        worksheet.cell(row=2, column=8).value = 'Retail + HNI (client)'
        worksheet.cell(row=2, column=8).alignment = Alignment(horizontal='center')
        worksheet.cell(row=2, column=8).fill = PatternFill(start_color="DDD9C4", fill_type="solid")
        worksheet.merge_cells('K2:M2')
        worksheet.cell(row=2, column=11).value = 'Proprietary'
        worksheet.cell(row=2, column=11).alignment = Alignment(horizontal='center')
        worksheet.cell(row=2, column=11).fill = PatternFill(start_color="FCD5B4", fill_type="solid")
        worksheet.merge_cells('N2:P2')
        worksheet.cell(row=2, column=14).value = 'Bank + NRI + others'
        worksheet.cell(row=2, column=14).alignment = Alignment(horizontal='center')
        worksheet.cell(row=2, column=14).fill = PatternFill(start_color="CC5CDA", fill_type="solid")
        #
        worksheet.cell(row=3, column=2).value = 'Buy'
        worksheet.cell(row=3, column=3).value = 'Sell'
        worksheet.cell(row=3, column=4).value = 'Net'

        worksheet.cell(row=3, column=5).value = 'Buy'
        worksheet.cell(row=3, column=6).value = 'Sell'
        worksheet.cell(row=3, column=7).value = 'Net'

        worksheet.cell(row=3, column=8).value = 'Buy'
        worksheet.cell(row=3, column=9).value = 'Sell'
        worksheet.cell(row=3, column=10).value = 'Net'

        worksheet.cell(row=3, column=11).value = 'Buy'
        worksheet.cell(row=3, column=12).value = 'Sell'
        worksheet.cell(row=3, column=13).value = 'Net'

        worksheet.cell(row=3, column=14).value = 'Buy'
        worksheet.cell(row=3, column=15).value = 'Sell'
        worksheet.cell(row=3, column=16).value = 'Net'

        for i in {3, 4}:
            for colNum in range(2, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).alignment = Alignment(wrap_text=True, horizontal='center')
                # worksheet.cell(row=i, column=colNum).style = headingStyle
        for i, column_cells in enumerate(worksheet.columns, start=2):
            worksheet.column_dimensions[get_column_letter(i)].width = 12
        for i in range(row_start, len(df) + row_start):
            for colNum in range(2, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[38]
        for i in range(2, len(df) + row_start):
            for colNum in {2, 5, 8, 11, 14, 17}:
                worksheet.cell(row=i, column=colNum).border = thin_border

    if sheet_name in {Participant_Interest_Sheet, Participant_Volumes_Sheet}:
        for i in range(row_start, len(df) + row_start):
            for colNum in range(3, worksheet.max_column + 1):
                worksheet.cell(row=i, column=colNum).number_format = BUILTIN_FORMATS[3]
    if sheet_name == FIIDerivativeSheet:
        for i in range(row_start, len(df) + row_start):
            worksheet.cell(row=i, column=2).style = 'Comma [0]'
            worksheet.cell(row=i, column=3).style = 'Comma [0]'
            worksheet.cell(row=i, column=4).style = 'Comma [0]'
            worksheet.cell(row=i, column=5).style = 'Comma [0]'
            worksheet.cell(row=i, column=6).style = 'Comma [0]'
            worksheet.cell(row=i, column=7).style = 'Comma [0]'
            worksheet.cell(row=i, column=8).style = 'Comma [0]'
            worksheet.cell(row=i, column=9).style = 'Comma [0]'
            worksheet.cell(row=i, column=10).style = 'Comma [0]'

    # if os.path.basename(filename) == final_inside_trading_file:
    #     for i in range(row_start, len(df) + row_start):
    #         for colNum in {8, 9, 10}:
    #             worksheet.cell(row=i, column=colNum).number_format = FORMAT_PERCENT_COLOR
    # save the workbook
    writer.save()


def get_page(currURL, type='NSE'):
    try:
        # print(indentation + "Requesting ")
        r = get_request_page(currURL, type)
        return r.text
    except Exception as e:
        print("Error in get URL " + str(e))
    return ""


def get_request_boot():
    global final_app_id_data
    global final_app_it_data
    currURL = "https://www.nseindia.com/companies-listing/corporate-filings-insider-trading"
    host = "www.nseindia.com"
    headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
               'Accept-Encoding': 'gzip, deflate, br',
               'Host': host,
               'Accept-Language': 'en-US,en;q=0.5',
               'Connection': 'keep-alive',
               'Upgrade-Insecure-Requests': '1',
               'DNT': '1',
               'TE': 'Trailers',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:80.0) Gecko/20100101 Firefox/80.0',
               'X-Requested-With': 'XMLHttpRequest'
               }
    try:
        response_data = requests.get(currURL, headers=headers)
        final_app_id_data = response_data.cookies['nseappid']
        final_app_it_data = response_data.cookies['nsit']

    except Exception as e:
        print("Error in get ID data " + str(e))
    return ""


def get_request_page(currURL, type='NSE'):
    headers = ""

    if type == "NEWNSE":
        host = "www.nseindia.com"
        headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Host': host,
                   'Accept-Language': 'en-US,en;q=0.5',
                   # 'Cookie': 'nsit=4zczKK0lbkTusHLjrgpuiG2P; nseappid=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJhcGkubnNlIiwiYXVkIjoiYXBpLm5zZSIsImlhdCI6MTYzMTE1ODI5NiwiZXhwIjoxNjMxMTYxODk2fQ._8a5psTHFgFM4nqqi5OEYWLtaePwmb6TL-Q7xWPPtG0;',
                   'Cookie': 'nsit={0}; nseappid={1};'.format(final_app_it_data, final_app_id_data),
                   'Connection': 'keep-alive',
                   'Upgrade-Insecure-Requests': '1',
                   'DNT': '1',
                   'TE': 'Trailers',
                   # "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-insider-trading",
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:80.0) Gecko/20100101 Firefox/80.0',
                   'X-Requested-With': 'XMLHttpRequest'
                   }

    if type == "NSE":
        headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7',
                   'Cache-Control': 'max-age=0',
                   'Connection': 'keep-alive',
                   'Upgrade-Insecure-Requests': '1',
                   'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36',
                   'X-Requested-With': 'XMLHttpRequest'
                   }

    else:
        if type == 'BSE':
            headers = {
                'Content-Type': 'text/html; charset=iso-8859-1',
                'Host': 'www.bseindia.com',
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip',
                'DNT': '1',
                'Connection': 'close'
            }
    try:
        response_data = requests.get(currURL, headers=headers)
        return response_data
    except Exception as e:
        print("Error in get URL " + str(e))


def write_to_file(filename, data):
    try:
        with open(filename, "w", encoding="utf-8") as file:
            file.write(str(data))
    except Exception as e:
        print("Error in Write " + str(e))


def read_from_file(filename):
    try:
        f = codecs.open(filename, 'r')
        readHtml = f.read()
        return readHtml
    except Exception as e:
        print("Error in Read " + str(e))


def write_to_download_file(filename, data):
    try:
        with open(filename, "wb") as file:
            file.write(data)
    except Exception as e:
        print("Error in Write " + str(e))


def get_download_page(currURL, type='NSE'):
    def get_filename_from_cd(cd):
        """
        Get filename from content-disposition
        """
        if not cd:
            return None
        fname = re.findall('filename=(.+)', cd)
        if len(fname) == 0:
            return None
        return fname[0]

    try:
        r = get_request_page(currURL, type='NEWNSE')
        content_disp = r.headers.get('Content-disposition')
        if content_disp:
            filename = get_filename_from_cd(r.headers.get('content-disposition'))
        return r.content, filename
    except Exception as e:
        print("Error in get donload " + str(e))
    return "", ""


class FIIDerivative(object):

    def __init__(self, day):
        self.date = day
        self.Index_Futures_BUY = 0
        self.Index_Futures_SELL = 0
        self.Index_Options_BUY = 0
        self.Index_Options_SELL = 0
        self.Stock_Futures_BUY = 0
        self.Stock_Futures_SELL = 0
        self.Stock_Options_BUY = 0
        self.Stock_Options_SELL = 0

    def setValue(self, derivativeName, buy, sell):
        derivativeName = derivativeName.replace(" ", "_")
        buyAttr = derivativeName + "_BUY"
        sellAttr = derivativeName + "_SELL"
        self.__setattr__(buyAttr, buy)
        self.__setattr__(sellAttr, sell)

    def getValue(self):
        return [
            self.date,
            self.Index_Futures_BUY,
            self.Index_Options_BUY,
            self.Stock_Futures_BUY,
            self.Stock_Options_BUY,
            self.Index_Futures_SELL,
            self.Index_Options_SELL,
            self.Stock_Futures_SELL,
            self.Stock_Options_SELL,
            0,
            0]


def read_final_file(final_file_name, sheet_name, create_sheet=True, **to_excel_kwargs):
    book = openpyxl.load_workbook(final_file_name)
    df = pd.DataFrame()
    if not sheet_name in book.sheetnames:
        if create_sheet:
            book.create_sheet(sheet_name)
            book.save(final_file_name)
    else:
        df = pd.read_excel(final_file_name, sheet_name=sheet_name, **to_excel_kwargs)
    return df


# ----------------------------------------------------------------------

def get_FII_Derivative():
    url = "https://www.fpi.nsdl.co.in/web/Reports/Latest.aspx"
    out_html_file = support_path + "FII_Derivative.html"
    try:
        html_data = get_page(url);
        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find_all("table", {"class": "tbls01"})[1]
        table_df = pd.read_html(str(data), header=3)[0]
        table_df.drop(table_df.tail(2).index, inplace=True)
        todayFIIDerivative = FIIDerivative(table_df.iloc[0][0])
        for index, row in table_df.iterrows():
            todayFIIDerivative.setValue(row[1], float(row[2].strip().strip("'")), float(row[4].strip().strip("'")))
        return todayFIIDerivative.getValue()
    except Exception as e:
        print(e)
        exit("ERROR FII")


def write_FII_final_file(final_df, final_excel_name, FIIDerivativeSheet, dataRow):
    if dataRow:
        today_df = pd.DataFrame([pd.Series(dataRow, index=['Date',
                                                           'Index Futures - BUY', 'Index Options - BUY',
                                                           'Stock Futures - BUY', 'Stock Options - BUY',
                                                           'Index Futures - SELL', 'Index Options - SELL',
                                                           'Stock Futures - SELL', 'Stock Options - SELL',
                                                           'Index Net', 'Stock Net'
                                                           ])])
    else:
        today_df = pd.DataFrame()
    if not final_df.empty:
        final_df['Date'] = final_df['Date'].astype(str)
        if 'Date' in final_df.columns:
            final_df.drop(final_df[final_df['Date'] == dataRow[0]].index, inplace=True)
        final_df = pd.concat([today_df, final_df], axis=0)
    else:
        final_df = today_df
    append_df_to_excel(final_excel_name, final_df, sheet_name=FIIDerivativeSheet, startrow=3, startcol=0,
                       truncate_sheet=False)


def process_fii_derivative(fileName):
    print("== ----------------------------FII Derivative--------------------------")
    dataRow = get_FII_Derivative()
    df = read_final_file(fileName, FIIDerivativeSheet, skiprows=topEmptyRows)
    write_FII_final_file(df, fileName, FIIDerivativeSheet, dataRow)


def get_cash_segment_FII():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/fiiEQ.htm"
    out_html_file = support_path + "fiiEQ.htm"
    try:
        html_data = get_page(url);
        write_to_file(out_html_file, html_data)
        html_data = read_from_file(out_html_file)
        soup = BeautifulSoup(html_data, 'lxml')
        fiitable = soup.find_all("table")[0]
        table_df = pd.read_html(str(fiitable), header=1)[0]
        todayFII = table_df.iloc[0][1]
        FII = {
            "Date": table_df.iloc[0][1],
            "FII Buy Value": table_df.iloc[0][2],
            "FII Sell Value": table_df.iloc[0][3],
            "FII Net Value": table_df.iloc[0][4]
        }
        FII_df = pd.DataFrame([pd.Series(FII)])
        return FII_df
    except Exception as e:
        print(e)
        exit("ERROR get_cash_segment_FII")


def get_cash_segment_DII():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/DiiEQ.htm"
    out_html_file = support_path + "DiiEQ.htm"
    try:
        html_data = get_page(url)
        write_to_file(out_html_file, html_data)
        html_data = read_from_file(out_html_file)
        soup = BeautifulSoup(html_data, 'lxml')
        fiitable = soup.find_all("table")[0]
        table_df = pd.read_html(str(fiitable), header=1)[0]
        todayFII = table_df.iloc[0][1]
        DII = {
            "Date": table_df.iloc[0][1],
            "DII Buy Value": table_df.iloc[0][2],
            "DII Sell Value": table_df.iloc[0][3],
            "DII Net Value": table_df.iloc[0][4]
        }
        DII_df = pd.DataFrame([pd.Series(DII)],
                              columns=['Date', 'DII Buy Value', 'DII Sell Value', 'DII Net Value'])
        return DII_df
    except Exception as e:
        print(e)
        exit("ERROR get_cash_segment_DII")


def get_cash_segment_FII_DII():
    try:
        FII_Data = get_cash_segment_FII()
        DII_Data = get_cash_segment_DII()
        # FII_DII_DF = pd.concat([FII_Data, DII_Data], axis=1)
        FII_DII_DF = pd.merge(left=FII_Data, right=DII_Data, on='Date')
        return FII_DII_DF
    except Exception as e:
        print(e)
        exit("ERROR get_cash_segment_FII_DII")


def get_NSE_turnover_data():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/HistoricalIndicesCapitalMarkets.htm"
    out_html_file = support_path + "HistoricalIndicesCapitalMarkets.htm"
    try:
        support_debug = False
        df = pd.DataFrame()
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find("td", text="Category-wise Turnover  (xls)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)
        xls_file = support_path + csv_filename
        if not support_debug:
            r = requests.get(csv_base_url)
            with open(xls_file, "wb") as file:
                file.write(r.content)

        df = pd.read_excel(xls_file, skiprows=2, skipfooter=1)
        return df
    except Exception as e:
        print(e)
        print(" No data for NSE, so Blank data")
        return df


def get_BSE_turnover_data():
    url = "https://www.bseindia.com/markets/equity/EQReports/categorywise_turnover.aspx"
    out_html_file = support_path + "categorywise_turnover.htm"
    table_df = pd.DataFrame()
    try:
        support_debug = False
        if os.path.isfile(out_html_file) and support_debug:
            content = read_from_file(out_html_file)
        else:
            content = get_page(url, "BSE")
            write_to_file(out_html_file, content)

        soup = BeautifulSoup(content, 'lxml')
        data = soup.find_all("div", {"id": "ContentPlaceHolder1_divOtherCat"})[0]
        data = data.find("table")
        for body in data("tbody"):
            body.unwrap()
        table_df = pd.read_html(str(data))[0]
        return table_df
    except Exception as e:
        print(" get_BSE_turnover_data -- " + e)
        return table_df


def get_formatted_date(date_string):
    try:
        formatter_string = "%b %d, %Y %H:%M:%S"
        datetime_object = datetime.strptime(date_string, formatter_string)
        date_object = datetime_object.date()
        date_mon = date_object.strftime("%d-%b")
        return date_mon
    except Exception as e:
        print(e)
        exit("ERROR Get Date format")


def get_cash_segment_Turnover():
    try:
        final_turnover_df = pd.DataFrame()

        NSE_Data = get_NSE_turnover_data()
        if not NSE_Data.empty:
            BSE_Data = get_BSE_turnover_data()
            nse_MarketDate = str(NSE_Data['Trade Date'][0])
            bse_formatted_date = datetime.strptime(nse_MarketDate, '%d-%b-%y').strftime('%d/%m/%Y')
            BSE_Data = BSE_Data.loc[BSE_Data[0] == bse_formatted_date]

            NSE_others_buy = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_OTHERS, NSE_BUY_COL_NAME].values[0])
            NSE_others_sell = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_OTHERS, NSE_SELL_COL_NAME].values[0])
            NSE_others_net = NSE_others_buy - NSE_others_sell
            NSE_protrade_buy = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_PROTRADE, NSE_BUY_COL_NAME].values[0])
            NSE_protrade_sell = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_PROTRADE, NSE_SELL_COL_NAME].values[0])
            NSE_protrade_net = NSE_protrade_buy - NSE_protrade_sell
            NSE_bank_buy = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_BANK, NSE_BUY_COL_NAME].values[0])
            NSE_bank_sell = float(NSE_Data.loc[NSE_Data[NSE_CATEGORY] == NSE_BANK, NSE_SELL_COL_NAME].values[0])
            NSE_bank_net = NSE_bank_buy - NSE_bank_sell
            BSE_clients_buy = float(BSE_Data[1].values[0])
            BSE_clients_sell = float(BSE_Data[2].values[0])
            BSE_clients_net = BSE_clients_buy - BSE_clients_sell
            BSE_NRI_buy = float(BSE_Data[4].values[0])
            BSE_NRI_sell = float(BSE_Data[5].values[0])
            BSE_NRI_net = BSE_NRI_buy - BSE_NRI_sell
            BSE_proprietary_buy = float(BSE_Data[7].values[0])
            BSE_proprietary_sell = float(BSE_Data[8].values[0])
            BSE_proprietary_net = BSE_proprietary_buy - BSE_proprietary_sell
            #
            retail_buy = NSE_others_buy + BSE_clients_buy
            retail_sell = NSE_others_sell + BSE_clients_sell
            retail_net = NSE_others_net + BSE_clients_net
            proprietary_buy = NSE_protrade_buy + BSE_proprietary_buy
            proprietary_sell = NSE_protrade_sell + BSE_proprietary_sell
            proprietary_net = NSE_protrade_net + BSE_proprietary_net
            bank_nri_buy = NSE_bank_buy + BSE_NRI_buy
            bank_nri_sell = NSE_bank_sell + BSE_NRI_sell
            bank_nri_net = NSE_bank_net + BSE_NRI_net
            column_list = ["Date", head_retail + ' Buy', head_retail + ' Sell', head_retail + ' Net',
                           head_proprietary + ' Buy', head_proprietary + ' Sell', head_proprietary + ' Net',
                           head_bank_nri + ' Buy', head_bank_nri + ' Sell', head_bank_nri + ' Net'
                           ]
            formatted_date = datetime.strptime(nse_MarketDate, '%d-%b-%y').strftime('%d-%b-%Y')
            final_turnover_df = pd.DataFrame([[formatted_date,
                                               retail_buy, retail_sell, retail_net,
                                               proprietary_buy, proprietary_sell, proprietary_net,
                                               bank_nri_buy, bank_nri_sell, bank_nri_net
                                               ]]
                                             , columns=column_list)
            final_turnover_df = final_turnover_df.set_index("Date")
            return final_turnover_df
        else:
            print("Returning Empty since, NSE has no data")
            return final_turnover_df
    except Exception as e:
        print(e)
        return final_turnover_df


def get_cash_segment_data():
    try:
        cash_segment_df = pd.DataFrame()
        currentMarketDate = ""
        FII_DII_DF = get_cash_segment_FII_DII()
        if not FII_DII_DF.empty:
            currentMarketDate = str(FII_DII_DF['Date'][0])
            turnover_DF = get_cash_segment_Turnover()
            if not turnover_DF.empty:
                cash_segment_df = pd.merge(left=FII_DII_DF, right=turnover_DF, on='Date', how="outer")
            else:
                print("No Cash Segment Data, so just FII DII only")
                cash_segment_df = FII_DII_DF

            return cash_segment_df, currentMarketDate
        else:
            return cash_segment_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_cash_segment_data")


def process_cash_segment(fileName):
    print("== ----------------------------Cash Segment--------------------------")
    cash_segment_df, currentMarketDate = get_cash_segment_data()
    currentMarketDate = str(currentMarketDate)
    final_df = read_final_file(fileName, CashSegmentSheet, skiprows=topEmptyRows)
    print("-----------------------------------=------------------------")
    if not (final_df.empty) and (column_CashSegmentDate in final_df.columns):
        final_df[column_CashSegmentDate] = final_df[column_CashSegmentDate].astype(str)
        final_df.drop(final_df[final_df['Date'] == str(currentMarketDate)].index, inplace=True)

    if not (final_df.empty):
        final_df = cash_segment_df.set_index('Date').combine_first(final_df.set_index('Date')).reset_index()
    else:
        final_df = cash_segment_df

    cash_segment_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[column_CashSegmentDate],
                         format='%d-%b-%Y', errors="ignore")).sort_values('x',
                                                                          ascending=False
                                                                          ).drop('x', 1)

    columnsTitles = ["Date",
                     "FII Buy Value", "FII Sell Value", "FII Net Value",
                     "DII Buy Value", "DII Sell Value", "DII Net Value",
                     head_retail + ' Buy', head_retail + ' Sell', head_retail + ' Net',
                     head_proprietary + ' Buy', head_proprietary + ' Sell', head_proprietary + ' Net',
                     head_bank_nri + ' Buy', head_bank_nri + ' Sell', head_bank_nri + ' Net'
                     ]
    final_df = final_df.reindex(columns=columnsTitles)
    write_common(final_df, fileName, CashSegmentSheet, cash_segment_df)


# ----------------------------------------
def get_nifty_top10_date(filename):
    try:
        date_string = re.findall(r"\_(.*?)\.", filename)[0]
        formatter_string = "%d%m%y"
        datetime_object = datetime.strptime(date_string, formatter_string)
        date_object = datetime_object.date()
        date_mon = date_object.strftime("%d-%b")
        # print("date_mon:", date_mon)
        return date_mon
    except Exception as e:
        print(e)
        exit("ERROR get_nifty_top10_date")


def get_nifty_top10_data():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/HistoricalIndicesCapitalMarkets.htm"
    out_html_file = support_path + "HistoricalIndicesCapitalMarkets.htm"

    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')

        data = soup.find("td", text="NIFTY 50  Top 10 Holdings  (csv)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)
        csv_data = get_page(csv_base_url)
        write_to_file(support_path + csv_filename, csv_data)
        csv_table_df = pd.read_csv(io.StringIO(csv_data), header=0, index_col="SYMBOL")

        day_mon = get_nifty_top10_date(csv_filename)
        return csv_table_df, day_mon

    except Exception as e:
        print(e)
        exit("ERROR FII")


def write_nifty_top10_final_file(df, final_excel_name, niftytop10Sheet, dataRow, day):
    column_symbol = "SYMBOL"
    column_security = "SECURITY"
    column_weightage = "WEIGHTAGE(%)"
    if not df.empty:
        if day in df.columns:
            print(indentation + " Data for {} - will be overwritten".format(day))
        else:
            df.insert(2, day, 0)
    else:
        df = pd.DataFrame(columns=(column_symbol, column_security))
    for row_label, row in dataRow.iterrows():
        row_already_exists = df.loc[df['SYMBOL'] == row_label].index
        if len(row_already_exists) > 0:
            df.loc[int(row_already_exists[0]), day] = float(row[column_weightage])
            df.loc[int(row_already_exists[0]), column_security] = str(row[column_security])
        else:
            df2 = pd.DataFrame(columns=df.columns)
            df2.loc[row_label, day] = float(row[column_weightage])
            df2.loc[row_label, column_symbol] = row_label
            df2.loc[row_label, column_security] = str(row[column_security])
            df = df.append(df2).fillna(0)
    append_df_to_excel(final_excel_name, df, sheet_name=niftytop10Sheet, startrow=3, startcol=0, truncate_sheet=True)


def process_nifty_top10():
    print("== ----------------------------Nifty Top10--------------------------")
    csv_data_df, day = get_nifty_top10_data()
    df = read_final_file(final_excel_name, NiftyTop10Sheet, skiprows=topEmptyRows)
    write_nifty_top10_final_file(df, final_excel_name, NiftyTop10Sheet, csv_data_df, day)


# --- --------------

def get_oi_spurts_data():
    url = "https://www1.nseindia.com/live_market/dynaContent/live_analysis/oi_spurts/topPositiveOIChangeData.json"
    out_html_file = support_path + "topPositiveOIChangeData.json"
    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
        oi_spurts_data = json.loads(html_data)
        currentMarketDate = oi_spurts_data['currentMarketDate']
        oi_spurts_df = pd.DataFrame(oi_spurts_data['data'])
        oi_spurts_df = oi_spurts_df.drop(['latestOI', 'valueInLakhs', 'prevOI', 'isFO'], axis=1)
        cols = ['oiChange', 'percOIchange', 'underlying', 'FUTVAL', 'OPTVAL', 'TOTVAL', 'OPVAL', 'volume']
        oi_spurts_df[cols] = oi_spurts_df[cols].apply(lambda x: x.str.replace(',', '').astype(float), axis=1)
        oi_spurts_df[column_oiSpurtsRank] = oi_spurts_df['TOTVAL'].rank(method='dense', ascending=False)
        oi_spurts_df.sort_values(by=[column_oiSpurtsRank], inplace=True)
        oi_spurts_df = oi_spurts_df.rename(columns={'symbol': 'Symbol',
                                                    'oiChange': 'Chg in OI',
                                                    'percOIchange': currentMarketDate,
                                                    'volume': 'Volume contracts',
                                                    'underlying': 'Underlying value (CM)',
                                                    'FUTVAL': 'Futures',
                                                    'OPTVAL': 'Options (Notional)',
                                                    'TOTVAL': 'Total',
                                                    'OPVAL': 'Options (Premium)'
                                                    })
        columnsTitles = ['Symbol', 'Chg in OI', 'Volume contracts', 'Underlying value (CM)',
                         'Futures', 'Options (Notional)', 'Total', 'Options (Premium)', column_oiSpurtsRank,
                         currentMarketDate]
        oi_spurts_df = oi_spurts_df.reindex(columns=columnsTitles)

        return oi_spurts_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR oi_spurts_data")


def write_oi_spurts_data(df, final_excel_name, OISpurtsSheet, data_df):
    df = df.append(data_df)
    append_df_to_excel(final_excel_name, df, sheet_name=OISpurtsSheet, startrow=0, startcol=0, truncate_sheet=True,
                       float_format="%0.2f")


def process_oi_spurts():
    print("== ----------------------------OI Spurts--------------------------")
    oi_data_df, currentMarketDate = get_oi_spurts_data()
    final_df = read_final_file(final_excel_name, OISpurtsSheet, skiprows=topEmptyRows)

    if not final_df.empty and (currentMarketDate in final_df.columns):
        final_df = final_df.drop([currentMarketDate], axis=1)

    if not final_df.empty:
        final_df = final_df.drop(
            ['Chg in OI', 'Volume contracts', 'Underlying value (CM)', 'Futures', 'Options (Notional)', 'Total',
             'Options (Premium)', column_oiSpurtsRank], axis=1)
        horizontal_stack = pd.merge(oi_data_df, final_df, on='Symbol', how='outer')
        oi_data_df = pd.DataFrame()
    else:
        horizontal_stack = final_df
    write_common(horizontal_stack, final_excel_name, OISpurtsSheet, oi_data_df)


def get_active_securities_date(date_string):
    try:
        formatter_string = "%b %d, %Y %H:%M:%S"
        datetime_object = datetime.strptime(date_string, formatter_string)
        date_object = datetime_object.date()
        date_mon = date_object.strftime("%d-%b")
        return date_mon
    except Exception as e:
        print(e)
        exit("ERROR Get Date format")


def get_active_securities_data(url, outfilename):
    try:

        if os.path.isfile(outfilename) and support_debug:
            html_data = read_from_file(outfilename)
        else:
            html_data = get_page(url)

        active_securities_data = json.loads(html_data)
        currentMarketDate = get_active_securities_date(active_securities_data['time'])
        active_securities_df = pd.DataFrame(active_securities_data['data'])
        active_securities_df = active_securities_df.drop(
            ['openPrice', 'highPrice', 'lowPrice', 'series', 'ltp', 'previousPrice', 'tradedQuantity',
             'lastCorpAnnouncement'], axis=1)
        active_securities_df.insert(0, column_ActiveSecurityDate, str(currentMarketDate))
        cols = ['netPrice', 'turnoverInLakhs']
        active_securities_df[cols] = active_securities_df[cols].apply(lambda x: x.str.replace(',', '').astype(float),
                                                                      axis=1)

        active_securities_df = active_securities_df.rename(columns={'symbol': 'Symbol',
                                                                    'netPrice': '% Change',
                                                                    'turnoverInLakhs': 'Value (in Lakhs)',
                                                                    'lastCorpAnnouncementDate': 'Latest Ex date'
                                                                    })
        return active_securities_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_active_securities_data")


def get_active_securities_data_value():
    url = "https://www1.nseindia.com/live_market/dynaContent/live_analysis/most_active/allTopValue1.json"
    out_html_file = support_path + "allTopValue1.json"
    try:
        active_securities_df, currentMarketDate = get_active_securities_data(url, out_html_file)
        return active_securities_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_active_securities_data_value")


def get_active_securities_data_volume():
    url = "https://www1.nseindia.com/live_market/dynaContent/live_analysis/most_active/allTopVolume1.json"
    out_html_file = support_path + "allTopVolume1.json"
    try:
        active_securities_df, currentMarketDate = get_active_securities_data(url, out_html_file)
        return active_securities_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_active_securities_data_volume")


def write_common(df, final_excel_name, ActiveSecuritiesSheet, data_df):
    if not data_df.empty:
        df = df.append(data_df)
    append_df_to_excel(final_excel_name, df, sheet_name=ActiveSecuritiesSheet, startrow=3, startcol=0,
                       truncate_sheet=False,
                       float_format="%0.2f")


def write_active_securities(df, final_excel_name, ActiveSecuritiesSheet, data_df):
    df = df.append(data_df)
    append_df_to_excel(final_excel_name, df, sheet_name=ActiveSecuritiesSheet, startrow=0, startcol=0,
                       truncate_sheet=True,
                       float_format="%0.2f")


def process_active_securities_value():
    print("== ----------------------------Active Securities Value--------------------------")
    col_prefix = 'Value '
    active_securities_df, currentMarketDate = get_active_securities_data_value()
    Lcolumn_ActiveSecurityDate = col_prefix + column_ActiveSecurityDate
    final_df = read_final_file(final_excel_name, ActiveSecuritiesValueSheet, skiprows=topEmptyRows)
    if not (final_df.empty) and (Lcolumn_ActiveSecurityDate in final_df.columns):
        final_df.drop(final_df[final_df[Lcolumn_ActiveSecurityDate] == currentMarketDate].index, inplace=True)
    active_securities_df = active_securities_df.add_prefix(col_prefix)
    final_df = active_securities_df.append(final_df)
    active_securities_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[col_prefix + column_ActiveSecurityDate],
                         format='%d-%b')).sort_values('x',
                                                      ascending=False
                                                      ).drop('x', 1)
    write_common(final_df, final_excel_name, ActiveSecuritiesValueSheet, active_securities_df)


def process_active_securities_volume():
    print("== ----------------------------Active Securities Volume--------------------------")
    col_prefix = 'Volume '
    active_securities_df, currentMarketDate = get_active_securities_data_volume()
    Lcolumn_ActiveSecurityDate = col_prefix + column_ActiveSecurityDate
    final_df = read_final_file(final_excel_name, ActiveSecuritiesVolumeSheet, skiprows=topEmptyRows)
    if not (final_df.empty) and (Lcolumn_ActiveSecurityDate in final_df.columns):
        final_df.drop(final_df[final_df[Lcolumn_ActiveSecurityDate] == currentMarketDate].index, inplace=True)
    active_securities_df = active_securities_df.add_prefix(col_prefix)
    final_df = active_securities_df.append(final_df)
    active_securities_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[col_prefix + column_ActiveSecurityDate],
                         format='%d-%b')).sort_values('x',
                                                      ascending=False
                                                      ).drop('x', 1)
    write_common(final_df, final_excel_name, ActiveSecuritiesVolumeSheet, active_securities_df)


def get_volume_gainers_data():
    url = "https://www1.nseindia.com/live_market/dynaContent/live_analysis/volume_spurts/volume_spurts.json"
    outfilename = support_path + "volume_spurts.json"
    try:
        if os.path.isfile(outfilename) and support_debug:
            html_data = read_from_file(outfilename)
        else:
            html_data = get_page(url)

        volume_gainers_data = json.loads(html_data)
        currentMarketDate = get_active_securities_date(volume_gainers_data['time'])
        volume_gainers_df = pd.DataFrame(volume_gainers_data['data'])
        volume_gainers_df = volume_gainers_df.drop(['name'], axis=1)
        volume_gainers_df.insert(0, column_ActiveSecurityDate, str(currentMarketDate))
        cols = ['turn_lkh', 'week1a', 'week2a', 'week1vc', 'week2vc', 'value', 'ltp', 'netpr']
        volume_gainers_df[cols] = volume_gainers_df[cols].apply(lambda x: x.str.replace(',', '').astype(float),
                                                                axis=1)
        volume_gainers_df.sort_values(by=['turn_lkh'], inplace=True)
        volume_gainers_df = volume_gainers_df.nlargest(15, 'turn_lkh')
        volume_gainers_df = volume_gainers_df.rename(columns={'sym': 'Symbol',
                                                              'ltp': 'LTP',
                                                              'turn_lkh': 'Today Volume',
                                                              'week1a': '1 Week Avg. Volume',
                                                              'week2a': '2 Weeks Avg. Volume',
                                                              'week1vc': '1 Week Change (No. of times)',
                                                              'week2vc': '2 Weeks Change (No. of times)',
                                                              'value': 'Turnovers (crs)',
                                                              'netpr': '% Chng'
                                                              })
        return volume_gainers_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_volume_gainers_data")


def process_volume_gainers(nifty_list):
    print("== ----------------------------Volume Gainers--------------------------")
    volume_gainers_df, currentMarketDate = get_volume_gainers_data()
    final_df = read_final_file(final_excel_name, Top25VolumeGainersSheet, skiprows=topEmptyRows)
    if not (final_df.empty) and (column_Top25VolumeGainersDate in final_df.columns):
        final_df.drop(final_df[final_df[column_Top25VolumeGainersDate] == currentMarketDate].index, inplace=True)
    final_df = volume_gainers_df.append(final_df)
    volume_gainers_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[column_ActiveSecurityDate],
                         format='%d-%b')).sort_values(['x', column_Top25VolumeGainersVolume],
                                                      ascending=False
                                                      ).drop('x', 1)
    if not final_df.empty and ('NIFTY' in final_df.columns):
        final_df = final_df.drop(['NIFTY'], axis=1)
    final_df.insert(1, 'NIFTY', final_df.Symbol.map(nifty_list.set_index('Symbol')['NIFTY'],
                                                    na_action='ignore').replace({np.NaN: 'No'}))
    write_common(final_df, final_excel_name, Top25VolumeGainersSheet, volume_gainers_df)


def get_top_value_data():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/HistoricalIndicesCapitalMarkets.htm"
    out_html_file = support_path + "HistoricalIndicesCapitalMarkets.htm"
    zip_extract_path = support_path + "top_value\\"
    column_TopValueTradedValue = 'TOTTRDVAL'

    column_TopValuePriceChange = 'Price Change'
    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find("td", text="Bhavcopy file  (csv)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)
        zip_file = support_path + csv_filename
        r = requests.get(csv_base_url)
        with open(zip_file, "wb") as code:
            code.write(r.content)
        zf = ZipFile(zip_file)
        zf.extractall(path=zip_extract_path)
        zf.close()
        current_csv_file = zip_extract_path + csv_filename.replace(".zip", "")
        csv_table_df = pd.read_csv(current_csv_file, header=0)
        csv_table_df = csv_table_df.drop(['SERIES', 'OPEN', 'HIGH', 'LOW', 'CLOSE', 'TOTTRDQTY', 'TOTALTRADES', 'ISIN'],
                                         axis=1)
        csv_table_df[column_TopValueRank] = csv_table_df[column_TopValueTradedValue].rank(ascending=0)
        csv_table_df.sort_values(by=[column_TopValueRank], inplace=True)
        csv_table_df = csv_table_df[(csv_table_df[column_TopValueRank] <= 15)]
        csv_table_df[column_TopValuePriceChange] = csv_table_df.apply(
            lambda row: (row.LAST - row.PREVCLOSE) / row.LAST * 100, axis=1)
        columnsTitles = ['TIMESTAMP', 'SYMBOL', 'TOTTRDVAL', column_TopValuePriceChange, 'LAST', 'PREVCLOSE',
                         column_TopValueRank]
        csv_table_df = csv_table_df.reindex(columns=columnsTitles)
        currentMarketDate = csv_table_df.iloc[0]['TIMESTAMP']
        csv_table_df = csv_table_df.rename(columns={'TIMESTAMP': column_TopValueDate, 'TOTTRDVAL': 'Total Trade Value'})
        csv_table_df = csv_table_df.loc[:, ~csv_table_df.columns.str.contains('^Unnamed')]
        return csv_table_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR FII")


def process_top_value():
    print("== ----------------------------Top Value--------------------------")
    top_value_df, currentMarketDate = get_top_value_data()
    #
    final_df = read_final_file(final_excel_name, TopValueSheet, skiprows=topEmptyRows)
    if not (final_df.empty) and (column_TopValueDate in final_df.columns):
        final_df.drop(final_df[final_df[column_TopValueDate] == currentMarketDate].index, inplace=True)
    final_df = top_value_df.append(final_df)
    top_value_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[column_TopValueDate],
                         format='%d-%b-%Y')).sort_values(['x', column_TopValueRank],
                                                         ascending=[False, True]
                                                         ).drop('x', 1)
    write_common(final_df, final_excel_name, TopValueSheet, top_value_df)


def get_MWPL_data():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/HistoricalIndicesDerivatives.htm"
    out_html_file = support_path + "HistoricalIndicesDerivatives.htm"
    zip_extract_path = support_path + "MWPL\\"
    try:

        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find("td", text="NSE Open Interest  (zip)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)

        zip_file = support_path + csv_filename

        r = requests.get(csv_base_url)
        with open(zip_file, "wb") as code:
            code.write(r.content)

        zf = ZipFile(zip_file)
        zf.extractall(path=zip_extract_path)
        zf.close()
        current_csv_file = zip_extract_path + csv_filename.replace(".zip", ".csv")
        csv_table_df = pd.read_csv(current_csv_file, header=0)
        csv_table_df.rename(columns=lambda x: x.strip(), inplace=True)
        currentMarketDate = csv_table_df.iloc[0]['Date']
        csv_table_df = csv_table_df.drop(['ISIN', 'Scrip Name', 'Date'], axis=1)

        csv_table_df[currentMarketDate] = csv_table_df.apply(
            lambda row: (row['NSE Open Interest'] / row['MWPL']) * 100, axis=1)
        return csv_table_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR MWPL")


def process_MWPL():
    print("== ----------------------------MWPL--------------------------")
    MWPL_df, currentMarketDate = get_MWPL_data()
    final_df = read_final_file(final_excel_name, MWPLSheet, skiprows=topEmptyRows)

    if not final_df.empty:
        if currentMarketDate in final_df.columns:
            final_df = final_df.drop([currentMarketDate], axis=1)
        final_df = final_df.drop(['MWPL', 'NSE Open Interest'], axis=1)
        horizontal_stack = pd.merge(MWPL_df, final_df, on='NSE Symbol', how='outer')
        MWPL_df = pd.DataFrame()
    else:
        horizontal_stack = final_df
    write_common(horizontal_stack, final_excel_name, MWPLSheet, MWPL_df)


def string2float(num):
    num = str(num).replace(' ', '').replace(',', '').replace("−", "-")
    return float(num)


def convert2float(val):
    try:
        return float(val)
    except ValueError:
        return np.nan


def get_nifty_index():
    url = "https://www1.nseindia.com/content/indices/ind_nifty50list.csv"
    out_html_file = support_path + "ind_nifty50list.csv"
    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            if not support_debug:
                write_to_file(out_html_file, html_data)
        csv_table_df = pd.read_csv(io.StringIO(html_data), header=0)
        csv_table_df.rename(columns=lambda x: x.strip(), inplace=True)
        csv_table_df['NIFTY'] = 'Yes'
        csv_table_df = csv_table_df.drop(['Company Name', 'ISIN Code', 'Series', 'Industry'], axis=1)
        return csv_table_df
    except Exception as e:
        print(e)
        exit("indexvalues_arr FII")


def process_nifty_index():
    dataRow = get_nifty_index()
    return dataRow


def get_FNO_data():
    try:
        url = "https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/foSecStockWatch.json"
        outfilename = support_path + "foSecStockWatch.json"

        if os.path.isfile(outfilename) and support_debug:
            html_data = read_from_file(outfilename)
        else:
            html_data = get_page(url)
            if support_debug:
                write_to_file(outfilename, html_data)

        FNO_data = json.loads(html_data)
        currentMarketDate = get_active_securities_date(FNO_data['time'])

        FNO_df = pd.DataFrame(FNO_data['data'])
        FNO_df = FNO_df.drop(
            ['open', 'high', 'low', 'trdVol', 'ptsC', 'trdVolM', 'mVal', 'wkhicm_adj', 'wklocm_adj', 'xDt', 'cAct'],
            axis=1)
        cols = ['per', 'ltP', 'ntP', 'wkhi', 'wklo']
        FNO_df[cols] = FNO_df[cols].apply(lambda x: x.str.replace(',', '').astype(float), axis=1)

        cols = ['mPC']
        FNO_df[cols] = FNO_df[cols].apply(lambda x: convert2float(x), axis=1)
        cols = ['yPC']
        FNO_df[cols] = FNO_df[cols].apply(lambda x: convert2float(x), axis=1)

        FNO_df = FNO_df.rename(columns={'per': '% Chng',
                                        'ltP': 'ltP',
                                        'ntP': 'Turnover (crs.)',
                                        'wkhi': '52w H',
                                        'wklo': '52w L',
                                        'yPC': '365 % chng',
                                        'mPC': currentMarketDate
                                        })
        return FNO_df, currentMarketDate

    except Exception as e:
        print(e)
        exit("ERROR get_FNO_data")


def process_FNO(nifty_list):
    print("== ----------------------------FNO Process--------------------------")

    FNO_df, currentMarketDate = get_FNO_data()
    final_df = read_final_file(final_excel_name, FNOSheet, skiprows=topEmptyRows)
    if not final_df.empty and (currentMarketDate in final_df.columns):
        final_df = final_df.drop([currentMarketDate], axis=1)
    if not final_df.empty:
        final_df = final_df.drop(
            ['% Chng', 'ltP', 'Turnover (crs.)', '52w H', '52w L', '365 % chng'], axis=1, errors='ignore')
        horizontal_stack = pd.merge(FNO_df, final_df, on='symbol', how='outer')
        FNO_df = pd.DataFrame()
    else:
        horizontal_stack = final_df
    if not horizontal_stack.empty and ('NIFTY' in horizontal_stack.columns):
        horizontal_stack = horizontal_stack.drop(['NIFTY'], axis=1)
    horizontal_stack.insert(1, 'NIFTY', horizontal_stack.symbol.map(nifty_list.set_index('Symbol')['NIFTY'],
                                                                    na_action='ignore').replace({np.NaN: 'No'}))
    write_common(horizontal_stack, final_excel_name, FNOSheet, FNO_df)


def get_equities_stock_watch_index():
    url = "https://www1.nseindia.com/live_market/dynaContent/live_watch/equities_stock_watch.htm"
    out_html_file = support_path + "equities_stock_watch.htm"
    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            if not support_debug:
                write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find_all("select", {"name": "bankNiftySelect"})[0]
        options = data.findAll('option')
        indexvalues_arr = {}
        for option in options:
            indexvalues_arr[option.text.strip()] = option["value"]
        return indexvalues_arr
    except Exception as e:
        print(e)
        exit("indexvalues_arr FII")


def process_equities_stock_watch_index():
    dataRow = get_equities_stock_watch_index()
    return dataRow


def get_equities_stock_watch_data(current_index):
    try:
        url = "https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/{0}StockWatch.json".format(
            current_index)
        outfilename = support_path + "{0}StockWatch.json".format(current_index)

        if os.path.isfile(outfilename) and support_debug:
            html_data = read_from_file(outfilename)
        else:
            html_data = get_page(url)
            if support_debug:
                write_to_file(outfilename, html_data)

        equities_data = json.loads(html_data)
        currentMarketDate = get_active_securities_date(equities_data['time'])

        current_index_data = pd.DataFrame(equities_data['data'])
        current_index_data = current_index_data.drop(
            ['open', 'high', 'low', 'ltP', 'ptsC', 'trdVolM', 'mVal', 'wkhicm_adj', 'wklocm_adj', 'xDt', 'cAct',
             'previousClose', 'trdVol', 'ntP', 'mVal', 'dayEndClose', 'iislPtsChange', 'iislPercChange'], axis=1)
        cols = ['per', 'wkhi', 'wklo']
        current_index_data[cols] = current_index_data[cols].apply(lambda x: x.str.replace(',', '').astype(float),
                                                                  axis=1)
        cols = ['mPC']
        current_index_data[cols] = current_index_data[cols].apply(lambda x: convert2float(x), axis=1)
        cols = ['yPC']
        current_index_data[cols] = current_index_data[cols].apply(lambda x: convert2float(x), axis=1)
        current_index_data = current_index_data.rename(columns={'per': '% Chng',
                                                                'wkhi': '52w H',
                                                                'wklo': '52w L',
                                                                'yPC': '365 % chng',
                                                                'mPC': currentMarketDate
                                                                })

        current_index_summary_df = pd.DataFrame(equities_data['latestData'])
        current_index_summary_df = current_index_summary_df.drop(['open', 'high', 'low', 'ltp', 'ch'], axis=1)
        cols = ['per', 'yHigh', 'yLow', 'yCls', 'mCls']
        current_index_summary_df[cols] = current_index_summary_df[cols].apply(
            lambda x: x.str.replace(',', '').astype(float),
            axis=1)

        current_index_summary_df = current_index_summary_df.rename(columns={'yHigh': '52w High',
                                                                            'yLow': '52w Low',
                                                                            'mCls': currentMarketDate,
                                                                            'yCls': '365 d % Chng',
                                                                            'per': '% Chng',
                                                                            'indexName': 'index'
                                                                            })
        columnsTitles = ['index', '% Chng', '52w High', '52w Low', '365 d % Chng', currentMarketDate]
        current_index_summary_df = current_index_summary_df.reindex(columns=columnsTitles)

        return current_index_data, current_index_summary_df, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_equities_stock_watch_data")


def process_equities_stock_watch(sectorList_df):
    print("== ----------------------------equities_stock_watch--------------------------")
    dataIndexes = process_equities_stock_watch_index()

    stocksList = sectorList_df['STOCKS'].dropna().unique().tolist()
    today_df = pd.DataFrame()
    todays_summary_df = pd.DataFrame()
    selected_sectors = ["Nifty Auto", "Nifty Financial Services"]
    currentMarketDate = ""
    for index, row in sectorList_df.iterrows():
        current_sector = row['STOCKS']
        current_sector_stocks_flg = row['INDIVIDUAL']
        current_index_df, current_index_summary_df, currentMarketDate = get_equities_stock_watch_data(
            dataIndexes[current_sector])
        current_index_df.insert(0, column_EquitiesStockWatchSector, str(current_sector))
        if current_sector_stocks_flg and str(current_sector_stocks_flg) not in ('nan', 'NO'):
            today_df = pd.concat([today_df, current_index_df])
        todays_summary_df = pd.concat([todays_summary_df, current_index_summary_df])

    current_index_summary_df = pd.DataFrame()
    current_index_df = pd.DataFrame()

    final_df = read_final_file(final_excel_name, StocksSheet, skiprows=topEmptyRows)
    if not final_df.empty and (currentMarketDate in final_df.columns):
        final_df = final_df.drop([currentMarketDate], axis=1)
    if not final_df.empty:
        final_df = final_df.drop(
            ['% Chng', '52w H', '52w L', '365 % chng'], axis=1)
        horizontal_stack = pd.merge(today_df, final_df, on=['symbol', 'Sector'], how='outer')
        today_df = pd.DataFrame()
    else:
        horizontal_stack = final_df

    write_common(horizontal_stack, final_excel_name, StocksSheet, today_df)

    final_df = read_final_file(final_excel_name, IndexReturnSummarySheet, skiprows=topEmptyRows)
    if not final_df.empty and (currentMarketDate in final_df.columns):
        final_df = final_df.drop([currentMarketDate], axis=1)
    if not final_df.empty:
        final_df = final_df.drop(
            ['% Chng', '52w High', '52w Low', '365 d % Chng'], axis=1)
        horizontal_stack = pd.merge(todays_summary_df, final_df, on=['index'], how='outer')
        todays_summary_df = pd.DataFrame()
    else:
        horizontal_stack = final_df

    write_common(horizontal_stack, final_excel_name, IndexReturnSummarySheet, todays_summary_df)


def get_index_return_data():
    try:
        url = "https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/liveIndexWatchData.json"
        outfilename = support_path + "liveIndexWatchData.json"

        if os.path.isfile(outfilename) and support_debug:
            html_data = read_from_file(outfilename)
        else:
            html_data = get_page(url)
            if support_debug:
                write_to_file(outfilename, html_data)

        index_return_data = json.loads(html_data)
        currentMarketDate = get_active_securities_date(index_return_data['data'][0]['timeVal'])

        current_index_data = pd.DataFrame(index_return_data['data'])
        current_index_data = current_index_data.drop(
            ['timeVal', 'previousClose', 'low', 'open', 'high', 'indexOrder'], axis=1)
        cols = ['yearHigh', 'yearLow', "percChange", "last"]
        current_index_data[cols] = current_index_data[cols].apply(lambda x: x.str.replace(',', '').astype(float),
                                                                  axis=1)
        current_index_data = current_index_data.rename(columns={'yearHigh': '52w High',
                                                                'yearLow': '52w Low',
                                                                'percChange': currentMarketDate,
                                                                'last': 'Current',
                                                                'indexName': 'index'
                                                                })
        columnsTitles = ['index', 'Current', '52w High', '52w Low', currentMarketDate]
        current_index_data = current_index_data.reindex(columns=columnsTitles)
        return current_index_data, currentMarketDate
    except Exception as e:
        print(e)
        exit("ERROR get_index_return_data")


def process_index_return(sectorList_df):
    print("== ----------------------------index_return--------------------------")

    current_index_df, currentMarketDate = get_index_return_data()
    sectors = sectorList_df['SECTOR']

    final_df = read_final_file(final_excel_name, IndexReturnSheet, skiprows=topEmptyRows)
    if not final_df.empty and (currentMarketDate in final_df.columns):
        final_df = final_df.drop([currentMarketDate], axis=1)
    if not final_df.empty:
        final_df = final_df.drop(
            ['Current', '52w High', '52w Low'], axis=1)
        horizontal_stack = pd.merge(current_index_df, final_df, on=['index'], how='outer')
        current_index_df = pd.DataFrame()
    else:
        horizontal_stack = final_df
    horizontal_stack = pd.merge(sectors, horizontal_stack, left_on='SECTOR', right_on='index')
    if not horizontal_stack.empty:
        horizontal_stack = horizontal_stack.drop(
            ['SECTOR'], axis=1)
    write_common(horizontal_stack, final_excel_name, IndexReturnSheet, current_index_df)


def sector_list():
    sectorList_df = pd.DataFrame()
    try:
        sectorList_df = read_final_file(final_excel_name, IndexReturnSetupSheet, skiprows=0)
        sectorList_df = sectorList_df.fillna("")
        return sectorList_df
    except Exception as e:
        print(e.strerror)
        return sectorList_df


def get_particpant_data_date(filename):
    try:
        date_string = re.findall(r"[^._]+(?=[^_]*$)", filename)[0]
        formatter_string = "%d%m%Y"
        datetime_object = datetime.strptime(date_string, formatter_string)
        date_object = datetime_object.date()
        date_mon = date_object.strftime("%d-%b-%Y")
        return date_mon
    except Exception as e:
        print(e)
        exit("ERROR get_particpant_data_date")


def get_participant_data():
    url = "https://www1.nseindia.com/products/dynaContent/equities/equities/htms/HistoricalIndicesDerivatives.htm"
    out_html_file = support_path + "HistoricalIndicesDerivatives.htm"

    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url)
            write_to_file(out_html_file, html_data)

        soup = BeautifulSoup(html_data, 'lxml')
        data = soup.find("td", text="Participant wise Open Interest  (csv)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)

        csv_data = get_page(csv_base_url)
        write_to_file(support_path + csv_filename, csv_data)
        open_interest_csv_table_df = pd.read_csv(io.StringIO(csv_data), skiprows=1, header=0,
                                                 skip_blank_lines=True, skipinitialspace=True)[:-1]
        open_interest_day_mon = get_particpant_data_date(csv_filename)
        if column_participant_date in open_interest_csv_table_df.columns:
            print(indentation + "For {}".format(open_interest_day_mon))
        else:
            open_interest_csv_table_df.insert(0, column_participant_date, open_interest_day_mon)
        open_interest_csv_table_df.rename(columns=lambda x: x.strip(), inplace=True)

        data = soup.find("td", text="Participant wise Trading Volumes  (csv)").find('a', {'href': True})['href']
        csv_base_url = urllib.parse.urljoin(base_url, data)
        csv_filename = os.path.basename(urlparse(csv_base_url).path)

        csv_data = get_page(csv_base_url)
        write_to_file(support_path + csv_filename, csv_data)
        str_data = io.StringIO(csv_data)
        trading_volume_csv_table_df = pd.read_csv(str_data, header=1,
                                                  skip_blank_lines=True, skipinitialspace=True)[:-1]
        trading_volume_day_mon = get_particpant_data_date(csv_filename)
        if column_participant_date in trading_volume_csv_table_df.columns:
            print(indentation + " Data for {}".format(trading_volume_day_mon))
        else:
            trading_volume_csv_table_df.insert(0, column_participant_date, trading_volume_day_mon)
        trading_volume_csv_table_df.rename(columns=lambda x: x.strip(), inplace=True)

        return open_interest_csv_table_df, open_interest_day_mon, trading_volume_csv_table_df, trading_volume_day_mon

    except Exception as e:
        print(e)
        exit("ERROR get_participant_data")


def process_participant(current_market_date, csv_data_df, participant_sheet_name):
    print("== ----------------------------Participant--------------------------")
    final_df = read_final_file(final_CashSegment_file_name, participant_sheet_name, skiprows=topEmptyRows)
    if not final_df.empty and (column_participant_date in final_df.columns):
        final_df.drop(final_df[final_df[column_participant_date] == current_market_date].index, inplace=True)
    final_df = csv_data_df.append(final_df)
    csv_data_df = pd.DataFrame()
    final_df = final_df.assign(
        x=pd.to_datetime(final_df[column_participant_date],
                         format='%d-%b-%Y')).sort_values(['x'],
                                                         ascending=False
                                                         ).drop('x', 1)
    write_common(final_df, final_CashSegment_file_name, participant_sheet_name, csv_data_df)


def process_participant_data():
    try:
        open_interest_df, open_interest_day_mon, trading_volume_df, trading_volume_day_mon = get_participant_data()
        process_participant(open_interest_day_mon, open_interest_df, Participant_Interest_Sheet)
        process_participant(trading_volume_day_mon, trading_volume_df, Participant_Volumes_Sheet)

    except Exception as e:
        print(e.strerror)


def get_file_trading_data(url, filename):
    out_html_file = support_path + filename
    support_debug = False
    try:
        if os.path.isfile(out_html_file) and support_debug:
            print(out_html_file)
            html_data = read_from_file(out_html_file)
            print(indentation + " **************** Local Reading")
        else:
            html_data = get_page(url, 'NEWNSE')
            byte_string = html_data.encode(u'utf-8')
            try:
                with open(out_html_file, "wb") as file:
                    file.write(byte_string)
            except Exception as e:
                print("Error in Write " + e)
        insider_data = json.loads(html_data)
        return insider_data
    except Exception as e:
        print(e)
        exit("ERROR get_file_trading_data")


def get_last_value_data():
    url = "https://www.nseindia.com/api/merged-daily-reports?key=favCapital"
    out_html_file = support_path + "favCapital.json"
    zip_extract_path = support_path + "equities\\"
    csv_table_df = pd.DataFrame()
    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url, 'NEWNSE')
            write_to_file(out_html_file, html_data)
        favCapital_data = json.loads(html_data)
        zip_file_link = [x for x in favCapital_data if x['name'] == 'CM - Bhavcopy(csv)'][0]['link']

        csv_filename = os.path.basename(urlparse(zip_file_link).path)
        zip_file = support_path + csv_filename
        r = requests.get(zip_file_link)
        # print("I am here , 1647")
        with open(zip_file, "wb") as code:
            code.write(r.content)
        zf = ZipFile(zip_file)
        # print("I am here , 1651")
        zf.extractall(path=zip_extract_path)
        # print("I am here , 1653")
        zf.close()
        current_csv_file = zip_extract_path + csv_filename.replace(".zip", "")
        csv_table_df = pd.read_csv(current_csv_file, header=0)
        csv_table_df = csv_table_df[csv_table_df['SERIES'] == 'EQ']
        csv_table_df = csv_table_df[['SYMBOL', 'LAST']]
        csv_table_df = csv_table_df.rename(columns={'LAST': 'Current Price'})
        return csv_table_df
    except Exception as e:
        print(e)
        print("ERROR get_last_value_data")
        return csv_table_df


def get_eps_data(url):
    eps_data = ""
    try:
        if url:
            html_data = get_page(url)
            my_dict = xmltodict.parse(html_data, process_namespaces=False)
            obj_eps = my_dict['xbrli:xbrl']['in-bse-fin:BasicEarningsLossPerShareFromContinuingOperations']
            if type(obj_eps) is list:
                eps_data = my_dict['xbrli:xbrl']['in-bse-fin:BasicEarningsLossPerShareFromContinuingOperations'][0][
                    '#text']
            else:
                eps_data = my_dict['xbrli:xbrl']['in-bse-fin:BasicEarningsLossPerShareFromContinuingOperations'][
                    '#text']
        else:
            eps_data = ""
        return eps_data
    except Exception as e:
        # print(e)
        return eps_data


def process_insider_trading():
    print("== ----------------------------insider_trading--------------------------")
    insider_trading_df, to_date = get_insider_trading_data()

    insider_trading_df = insider_trading_df.drop_duplicates(subset="SYMBOL")
    inside_trading_data_sheet = str(to_date)
    append_df_to_excel(final_inside_trading_file_name, insider_trading_df, sheet_name=inside_trading_data_sheet,
                       startrow=3,
                       startcol=0,
                       truncate_sheet=True,
                       float_format="%0.2f")


def get_insider_trading_data():
    three_months_backdate = datetime.today() - pd.DateOffset(months=3)
    from_date = three_months_backdate.strftime("%d-%m-%Y")
    today = date.today()
    to_date = today.strftime("%d-%m-%Y")
    print("from_date =", from_date, "to_date =", to_date)
    get_csv = "&csv=true"
    previous_day_df = previous_day_insider_trading_data(today)

    columnsTitles = ['COMPANY', 'VALUE OF SECURITY(ACQUIRED) Cr', 'VALUE OF SECURITY(SELL) Cr',
                     'DII/FII/MF SELL', 'No of Securities BUY',
                     'Current Price', 'DATE OF ALLOTMENT/ACQUISITION TO', 'Avg Price', 'Jump in Price']

    previous_day_df = previous_day_df.drop(columnsTitles, axis=1, errors='ignore')

    url = "https://www.nseindia.com/api/corporates-pit?index=equities&from_date={0}&to_date={1}".format(
        from_date, to_date)
    out_html_file = support_path + "3-months-data.json"
    support_debug = False

    try:
        if os.path.isfile(out_html_file) and support_debug:
            html_data = read_from_file(out_html_file)
        else:
            html_data = get_page(url, 'NEWNSE')
            write_to_file(out_html_file, html_data)
        url = url + get_csv
        download_data, download_file = get_download_page(url, type='NEWNSE')
        write_to_download_file(support_path + download_file, download_data)
        insider_data = json.loads(html_data)
        insider_df = pd.DataFrame(insider_data['data'])
        insider_df = insider_df.drop(['anex', 'remarks', 'xbrl'], axis=1)
        personCategoryList = ['Promoter Group', 'Promoters', 'Key Managerial Personnel', 'Director']
        columns = ['symbol', 'company', 'acqMode', 'secVal', 'secAcq']
        secondset_columns = ['symbol', 'acqtoDt']
        key_field = "symbol"

        buysellacqModeList = ['Market Purchase', 'Market Sale']
        insider_buy_sell_data_df = insider_df[
            insider_df.personCategory.isin(personCategoryList) & insider_df.acqMode.isin(buysellacqModeList)]
        insider_buy_sell_df = insider_buy_sell_data_df.loc[:, columns]
        insider_buy_date_df = insider_buy_sell_data_df.loc[:, secondset_columns]

        insider_buy_date_df = insider_buy_date_df.groupby([key_field]).agg({'acqtoDt': ['max']})
        insider_buy_date_df.columns = ['DATE OF ALLOTMENT/ACQUISITION TO']

        insider_buy_date_df.reset_index(inplace=True)

        insider_buy_sell_df.loc[:, ['secVal']] = insider_buy_sell_df.loc[:, ['secVal']].apply(pd.to_numeric)
        insider_buy_sell_df.loc[:, ['secAcq']] = insider_buy_sell_df.loc[:, ['secAcq']].apply(pd.to_numeric)

        insider_buy_sell_df = pd.pivot_table(insider_buy_sell_df, index=['symbol', 'company'],
                                             values=["secVal", "secAcq"],
                                             columns=["acqMode"],
                                             aggfunc=[np.sum],
                                             fill_value=0)
        insider_buy_sell_df.columns = [f'{k}_{j}_{i}' for i, j, k in insider_buy_sell_df.columns]
        insider_buy_sell_df = insider_buy_sell_df[insider_buy_sell_df['Market Purchase_secVal_sum'] >= 8500000]
        insider_buy_sell_df = insider_buy_sell_df.drop(["Market Sale_secAcq_sum"], axis=1)

        insider_buy_sell_df.reset_index(inplace=True)
        horizontal_stack = pd.merge(insider_buy_sell_df, insider_buy_date_df, how='left', left_on='symbol',
                                    right_on='symbol').fillna("")
        promoter_holding = "https://www.nseindia.com/api/corporate-share-holdings-master?index=equities&from_date={0}&to_date={1}".format(
            from_date, to_date)
        promoter_holding_file = "promoter_holding.json"
        promoter_holding_data = get_file_trading_data(promoter_holding, promoter_holding_file)

        url = promoter_holding + get_csv
        download_data, download_file = get_download_page(url, type='NEWNSE')
        write_to_download_file(support_path + download_file, download_data)

        promoter_holding_df = pd.DataFrame(promoter_holding_data)
        if not promoter_holding_df.empty:
            promoter_holding_df = promoter_holding_df.set_index('symbol')
            promoter_holding_df = promoter_holding_df.rename(columns={'pr_and_prgrp': 'Promoter Holding'})
            promoter_holding_df = promoter_holding_df.drop(['desc', 'industry', 'name', 'date', 'xbrl', 'revisedData',
                                                            'recordId', 'cgTimeStamp', 'remarksWeb', 'countString',
                                                            'isin',
                                                            'public_val', 'underlyingDrs', 'employeeTrusts', 'index'
                                                            ], axis=1)
            horizontal_stack = pd.merge(horizontal_stack, promoter_holding_df, how='left', left_on='symbol',
                                        right_on='symbol').fillna("")

        else:
            horizontal_stack['Promoter Holding'] = ""
        horizontal_stack['Remarks1'] = ""
        horizontal_stack['Remarks2'] = ""

        for ind in horizontal_stack[(horizontal_stack['Promoter Holding']).isnull()].index:

            promoter_holding = "https://www.nseindia.com/api/corporate-share-holdings-master?index=equities&symbol={0}&issuer={1}".format(
                horizontal_stack['symbol'][ind], horizontal_stack['company'][ind])
            promoter_holding_file = horizontal_stack['symbol'][ind] + "promoter_holding.json"
            promoter_holding_data = get_file_trading_data(promoter_holding, promoter_holding_file)
            if promoter_holding_data:
                promoter_holding_df = pd.DataFrame(promoter_holding_data)
                promoter_holding_df = promoter_holding_df.set_index('symbol')
                promoter_holding_df = promoter_holding_df.rename(columns={'pr_and_prgrp': 'Promoter Holding'})
                promoter_holding_df.sort_values('date', ascending=False, inplace=True)
                promoter_holding_df = promoter_holding_df.head(1)
                promoter_holding_df = promoter_holding_df.drop(
                    ['desc', 'industry', 'name', 'date', 'xbrl', 'revisedData', 'recordId', 'cgTimeStamp', 'remarksWeb',
                     'countString', 'isin', 'public_val', 'underlyingDrs', 'employeeTrusts', 'index'
                     ], axis=1)

                horizontal_stack.loc[
                    (horizontal_stack['symbol'] == horizontal_stack['symbol'][ind]), ['Promoter Holding']] = \
                    promoter_holding_df.loc[horizontal_stack['symbol'][ind], 'Promoter Holding']

        pledge_url = "https://www.nseindia.com/api/corporate-pledgedata?index=equities&from_date={0}&to_date={1}".format(
            from_date, to_date)
        pledge_data_file = "pledge_data.json"
        pledge_data = get_file_trading_data(pledge_url, pledge_data_file)
        url = pledge_url + get_csv
        download_data, download_file = get_download_page(url, type='NEWNSE')
        write_to_download_file(support_path + download_file, download_data)

        pledge_df = pd.DataFrame(pledge_data['data'])

        pledge_df = pledge_df.rename(columns={'percPromoterShares': 'Pledge Data'})
        pledge_df = pledge_df.drop(
            ['shp', 'totIssuedShares', 'totPromoterHolding', 'percPromoterHolding',
             'totPublicHolding', 'totPromoterShares', 'percTotShares',
             'disclosureFromDate', 'numSharesPledged',
             'totDematShares', 'sharesCollateral', 'nbfcPromoShare', 'nbfcNonPromoShare',
             'percSharesPledged', 'broadcastDt', 'disclosureToDate', 'compBroadcastDate'
             ], axis=1)
        pledge_df['Pledge Data'] = pledge_df['Pledge Data'].str.strip()

        horizontal_stack = pd.merge(horizontal_stack, pledge_df, how='left', left_on='company',
                                    right_on='comName').fillna("")
        horizontal_stack = horizontal_stack.drop('comName', axis=1)

        FII_DII_MF_url = "https://www.nseindia.com/api/corporate-sast-reg29?index=equities&from_date={0}&to_date={1}".format(
            from_date, to_date)
        FII_data_file = "FII_DII_MF_data.json"
        FII_DII_MF_data = get_file_trading_data(FII_DII_MF_url, FII_data_file)

        url = FII_DII_MF_url + get_csv
        download_data, download_file = get_download_page(url, type='NEWNSE')
        write_to_download_file(support_path + download_file, download_data)

        FII_DII_MF_df = pd.DataFrame(FII_DII_MF_data['data'])
        FII_DII_MF_df = FII_DII_MF_df.rename(columns={'noOfShareSale': 'DII/FII/MF SELL'})
        FII_DII_MF_df = FII_DII_MF_df.drop(
            ['acquirerName', 'acquirerDate', 'noOfShareAcq', 'noOfShareAft',
             'regType', 'application_no', 'promoterType',
             'acqSaleType', 'acquisitionMode',
             'acqType', 'totAcqShare', 'totAcqDiluted', 'totSaleShare',
             'totSaleDiluted', 'totAftShare', 'totAftDiluted', 'remarks', 'attachement', 'time', 'company'
             ], axis=1)
        FII_DII_MF_df['timestamp'] = pd.to_datetime(FII_DII_MF_df.timestamp)
        FII_DII_MF_df = FII_DII_MF_df.sort_values(['symbol', 'timestamp'], ascending=[True, False]) \
            .groupby('symbol').head(1).drop('timestamp', 1)

        horizontal_stack = pd.merge(horizontal_stack, FII_DII_MF_df, how='left', left_on='symbol',
                                    right_on='symbol').fillna("")

        financial_result_url = "https://www.nseindia.com/api/corporates-financial-results?index=equities&period=Quarterly&from_date={0}&to_date={1}".format(
            from_date, to_date)
        financial_result_data_file = "financial_result_data.json"
        financial_result_data = get_file_trading_data(financial_result_url, financial_result_data_file)
        url = financial_result_url + get_csv
        download_data, download_file = get_download_page(url, type='NEWNSE')
        write_to_download_file(support_path + download_file, download_data)
        financial_result_df = pd.DataFrame(financial_result_data)
        financial_result_df['filingDate'] = pd.to_datetime(financial_result_df.filingDate)
        financial_result_df = financial_result_df.sort_values(
            ['symbol', 'filingDate', 'audited', 'consolidated'], ascending=[True, False, True, True]).groupby(
            'symbol').head(1).drop(['filingDate', 'audited', 'consolidated'], 1)
        financial_result_df = financial_result_df[['symbol', 'xbrl']]

        horizontal_stack = pd.merge(horizontal_stack, financial_result_df, how='left', left_on='symbol',
                                    right_on='symbol').fillna("")

        horizontal_stack['EPS'] = horizontal_stack.apply(lambda row: get_eps_data(row.xbrl), axis=1)
        horizontal_stack = horizontal_stack.drop('xbrl', axis=1)
        current_market_df = get_last_value_data()
        # current_market_df =[]
        if not current_market_df.empty:
            horizontal_stack = pd.merge(horizontal_stack, current_market_df, how='left', left_on='symbol',
                                        right_on='SYMBOL').fillna("")
            horizontal_stack = horizontal_stack.drop('SYMBOL', axis=1, errors='ignore')
        horizontal_stack = horizontal_stack.rename(
            columns={'Market Purchase_secVal_sum': 'VALUE OF SECURITY(ACQUIRED) Cr'
                , 'Market Sale_secVal_sum': 'VALUE OF SECURITY(SELL) Cr'
                , 'Market Purchase_secAcq_sum': 'No of Securities BUY'
                , 'symbol': 'SYMBOL'
                , 'company': 'COMPANY'
                     })

        horizontal_stack = horizontal_stack.replace('', np.nan)
        horizontal_stack = horizontal_stack.set_index('SYMBOL')
        if not previous_day_df.empty:
            previous_day_df = previous_day_df[previous_day_df['SYMBOL'].notna()]
            previous_day_df = previous_day_df.set_index('SYMBOL')
            horizontal_stack['Promoter Holding'].fillna(previous_day_df['Promoter Holding'], inplace=True)
            horizontal_stack['Pledge Data'].fillna(previous_day_df['Pledge Data'], inplace=True)
            horizontal_stack['EPS'].fillna(previous_day_df['EPS'], inplace=True)
            if 'Remarks1' in previous_day_df:
                horizontal_stack['Remarks1'].fillna(previous_day_df['Remarks1'], inplace=True)
            if 'Remarks1' in previous_day_df:
                horizontal_stack['Remarks2'].fillna(previous_day_df['Remarks2'], inplace=True)
        horizontal_stack.reset_index(inplace=True)
        horizontal_stack['Avg Price'] = horizontal_stack.apply(
            lambda row: (row['VALUE OF SECURITY(ACQUIRED) Cr'] / row['No of Securities BUY']), axis=1)
        horizontal_stack['Jump in Price'] = horizontal_stack.apply(
            lambda row: ((100 * (float(row['Current Price']) - float(row['Avg Price']))) / float(
                row['Avg Price'])) if (row['Current Price'] and row['Avg Price']) else '',
            axis=1)
        columnsTitles = [
            'COMPANY', 'SYMBOL'
            , 'VALUE OF SECURITY(ACQUIRED) Cr', 'VALUE OF SECURITY(SELL) Cr',
            'DII/FII/MF SELL', 'No of Securities BUY', 'Promoter Holding',
            'Current Price', 'Pledge Data', 'EPS', 'DATE OF ALLOTMENT/ACQUISITION TO', 'Avg Price', 'Jump in Price',
            'Remarks1', 'Remarks2']
        horizontal_stack = horizontal_stack.reindex(columns=columnsTitles)
        ##print(horizontal_stack)

        return horizontal_stack, to_date
    except Exception as e:
        print(e)
        exit("ERROR insider_df")


def previous_day_insider_trading_data(todays_date):
    # print("== ----------------------------get prev insider_trading--------------------------")
    check_day_limit = 3
    df = pd.DataFrame()
    for back_date in range(check_day_limit):
        current_backdate = todays_date - pd.DateOffset(days=back_date + 1)
        Previous_date = current_backdate.strftime("%d-%m-%Y")
        df = read_final_file(final_inside_trading_file_name, str(Previous_date), create_sheet=False,
                             skiprows=topEmptyRows)
        if not df.empty:
            break
    return df


def All_other_sheet():
    try:
        sectorList = sector_list()
        niftyStockList = process_nifty_index()
        process_nifty_top10()
        process_oi_spurts()
        process_active_securities_value()
        process_active_securities_volume()
        process_volume_gainers(niftyStockList)
        process_top_value()
        process_FNO(niftyStockList)
        process_equities_stock_watch(sectorList)
        process_index_return(sectorList)
        process_MWPL()

    except Exception as e:
        print(e)


def FII_cash_Segement_sheet(fileName):
    try:
        # second List
        process_fii_derivative(fileName)
        process_cash_segment(fileName)
        process_participant_data()
    except Exception as e:
        print(e)


if __name__ == "__main__":
    print("== ................................Stating................................")
    setup()
    # take_backup()
    # All_other_sheet()
    # FII_cash_Segement_sheet(final_CashSegment_file_name)
    process_insider_trading()

    print("== ................................Done................................")

